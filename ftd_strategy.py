"""
FTD (Follow-Through Day) Trading Strategy - Enhanced Exit Analysis

This strategy identifies oversold conditions using RSI, then identifies subsequent
FTD using high IBS (Intra-Bar Strength) value and high daily return (>1%).

TOP 5 ENTRY VARIATIONS (from prior analysis) - All with FTD criteria:
1. Entry_10: Trend Confirmation - RSI < 30 + FTD + Price above 200-day MA
2. Entry_15: Wide Bar - RSI < 30 + FTD + Large green bar (>1.5%)
3. Entry_7: Volume Spike - RSI < 30 + FTD + Volume > 1.5x average
4. Entry_8: Gap Up - RSI < 30 + FTD + Gap up > 0.5%
5. Entry_5: RSI < 40 + FTD (IBS > 0.5, any positive return)

EXIT VARIATIONS (9 total - removed Target):
RSI Threshold Variations (5):
- Exit_RSI_55: Exit when RSI > 55
- Exit_RSI_60: Exit when RSI > 60
- Exit_RSI_65: Exit when RSI > 65
- Exit_RSI_70: Exit when RSI > 70
- Exit_RSI_75: Exit when RSI > 75

New Exit Methodologies (4):
- Exit_Time_3: Hold for 3 bars
- Exit_Time_10: Hold for 10 bars
- Exit_RSI_Oversold: Exit when RSI goes back below 30
- Exit_MA_Cross: Exit when price crosses below 20-day MA

Total: 45 strategy variations (5 entries x 9 exits)
"""

import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings('ignore')

# Configuration
TICKER = 'SPY'
START_DATE = '2000-01-01'
END_DATE = '2024-12-31'


def calculate_rsi(prices, period=14):
    """Calculate RSI indicator"""
    delta = prices.diff()
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)
    
    avg_gain = gain.rolling(window=period, min_periods=period).mean()
    avg_loss = loss.rolling(window=period, min_periods=period).mean()
    
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi


def calculate_ibs(df):
    """Calculate Intra-Bar Strength: (close - low) / (high - low)"""
    ibs = (df['Close'] - df['Low']) / (df['High'] - df['Low'])
    return ibs


def calculate_daily_return(df):
    """Calculate daily return"""
    return df['Close'].pct_change()


def calculate_indicators(df):
    """Calculate all additional technical indicators"""
    # Moving averages
    df['MA_20'] = df['Close'].rolling(window=20).mean()
    df['MA_50'] = df['Close'].rolling(window=50).mean()
    df['MA_200'] = df['Close'].rolling(window=200).mean()
    
    # VWAP (approximate using typical price - cumulative calculation)
    df['Typical_Price'] = (df['High'] + df['Low'] + df['Close']) / 3
    df['Cum_Volume'] = df['Volume'].cumsum()
    df['Cum_TP_Volume'] = (df['Typical_Price'] * df['Volume']).cumsum()
    df['VWAP'] = df['Cum_TP_Volume'] / df['Cum_Volume']
    
    # Volume
    df['Volume_MA'] = df['Volume'].rolling(window=20).mean()
    df['Volume_Ratio'] = df['Volume'] / df['Volume_MA']
    
    # Gap calculation
    df['Gap'] = (df['Open'] - df['Close'].shift(1)) / df['Close'].shift(1)
    
    # Prior day data for swing high/low
    df['Prior_High'] = df['High'].shift(1)
    df['Prior_Low'] = df['Low'].shift(1)
    
    # Swing high (last 5 days)
    df['Swing_High'] = df['High'].rolling(window=5).max().shift(1)
    
    # Prior RSI
    df['Prior_RSI'] = df['RSI'].shift(1)
    
    return df


def download_spy_data():
    """Download SPY data from Yahoo Finance"""
    print(f"Downloading {TICKER} data...")
    df = yf.download(TICKER, start=START_DATE, end=END_DATE, progress=False)
    df = df.reset_index()
    
    # Flatten multi-level columns if present
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [col[0] if col[1] == '' or col[1] == TICKER else col[0] for col in df.columns]
    
    # Calculate base indicators
    df['RSI'] = calculate_rsi(df['Close'])
    df['IBS'] = calculate_ibs(df)
    df['Daily_Return'] = calculate_daily_return(df)
    
    # Calculate additional indicators
    df = calculate_indicators(df)
    
    print(f"Downloaded {len(df)} trading days of data")
    return df


# Entry condition functions
def entry_condition_1(df):
    """Entry 1: RSI < 30 (oversold) + FTD (IBS > 0.5 AND daily return > 1%)"""
    oversold = df['RSI'] < 30
    ftd = (df['IBS'] > 0.5) & (df['Daily_Return'] > 0.01)
    return oversold & ftd


def entry_condition_2(df):
    """Entry 2: RSI < 35 + FTD (IBS > 0.7 AND daily return > 1.5%)"""
    oversold = df['RSI'] < 35
    ftd = (df['IBS'] > 0.7) & (df['Daily_Return'] > 0.015)
    return oversold & ftd


def entry_condition_3(df):
    """Entry 3: RSI < 25 (more oversold) + FTD (IBS > 0.5 AND daily return > 1%)"""
    oversold = df['RSI'] < 25
    ftd = (df['IBS'] > 0.5) & (df['Daily_Return'] > 0.01)
    return oversold & ftd


def entry_condition_4(df):
    """Entry 4: RSI crosses above 30 (RSI reversal) + FTD"""
    rsi_cross = (df['Prior_RSI'] <= 30) & (df['RSI'] > 30)
    ftd = (df['IBS'] > 0.5) & (df['Daily_Return'] > 0.01)
    return rsi_cross & ftd


def entry_condition_5(df):
    """Entry 5: RSI < 40 + FTD (IBS > 0.5, any positive return)"""
    oversold = df['RSI'] < 40
    ftd = (df['IBS'] > 0.5) & (df['Daily_Return'] > 0)
    return oversold & ftd


# Additional Entry conditions (6-15) - Social Media strategies
def entry_condition_6(df):
    """Entry 6: MA Crossover - Price crosses above 50-day MA after oversold RSI"""
    rsi_oversold = df['RSI'] < 30
    ma_cross = df['Close'] > df['MA_50']
    prior_below = df['Close'].shift(1) <= df['MA_50'].shift(1)
    return rsi_oversold & ma_cross & prior_below


def entry_condition_7(df):
    """Entry 7: Volume Spike - Volume > 1.5x average after RSI oversold"""
    rsi_oversold = df['RSI'] < 30
    volume_spike = df['Volume_Ratio'] > 1.5
    return rsi_oversold & volume_spike


def entry_condition_8(df):
    """Entry 8: Gap Up - Gap up > 0.5% after RSI oversold"""
    rsi_oversold = df['RSI'] < 30
    gap_up = df['Gap'] > 0.005
    return rsi_oversold & gap_up


def entry_condition_9(df):
    """Entry 9: Support Bounce - Price bounces from daily VWAP after oversold"""
    rsi_oversold = df['RSI'] < 30
    # Close is above VWAP today but was below yesterday
    above_vwap = df['Close'] > df['VWAP']
    below_yesterday = df['Close'].shift(1) <= df['VWAP'].shift(1)
    return rsi_oversold & above_vwap & below_yesterday


def entry_condition_10(df):
    """Entry 10: Trend Confirmation - Price above 200-day MA after RSI oversold"""
    rsi_oversold = df['RSI'] < 30
    above_ma200 = df['Close'] > df['MA_200']
    return rsi_oversold & above_ma200


def entry_condition_11(df):
    """Entry 11: Momentum - RSI crosses above 40 with strong IBS (>0.7)"""
    rsi_cross = (df['Prior_RSI'] <= 40) & (df['RSI'] > 40)
    strong_ibs = df['IBS'] > 0.7
    return rsi_cross & strong_ibs


def entry_condition_12(df):
    """Entry 12: Pullback - Price retraces to 20-day MA after FTD"""
    ftd = (df['IBS'] > 0.5) & (df['Daily_Return'] > 0.01)
    # Price is at or near 20-day MA (within 1%)
    at_ma20 = abs(df['Close'] - df['MA_20']) / df['MA_20'] < 0.01
    return ftd & at_ma20


def entry_condition_13(df):
    """Entry 13: Swing High Break - Break of prior swing high after oversold"""
    rsi_oversold = df['RSI'] < 30
    break_swing = df['Close'] > df['Swing_High']
    return rsi_oversold & break_swing


def entry_condition_14(df):
    """Entry 14: VWAP Support - Price bounces from VWAP after oversold"""
    rsi_oversold = df['RSI'] < 30
    # Price is above VWAP today but was below yesterday (bounce)
    bounce = (df['Close'] > df['VWAP']) & (df['Close'].shift(1) <= df['VWAP'].shift(1))
    return rsi_oversold & bounce


def entry_condition_15(df):
    """Entry 15: Wide Bar - Large green bar (>1.5%) after RSI oversold"""
    rsi_oversold = df['RSI'] < 30
    wide_bar = df['Daily_Return'] > 0.015
    return rsi_oversold & wide_bar


# Exit condition functions
# RSI Threshold Variatons (5)
def exit_condition_rsi_55(df, threshold=55):
    """Exit: RSI reaches threshold"""
    return df['RSI'] > threshold


def exit_condition_rsi_60(df, threshold=60):
    """Exit: RSI reaches threshold"""
    return df['RSI'] > threshold


def exit_condition_rsi_65(df, threshold=65):
    """Exit: RSI reaches threshold"""
    return df['RSI'] > threshold


def exit_condition_rsi_70(df, threshold=70):
    """Exit: RSI reaches threshold"""
    return df['RSI'] > threshold


def exit_condition_rsi_75(df, threshold=75):
    """Exit: RSI reaches threshold"""
    return df['RSI'] > threshold


# New Exit Methodologies (4) - Removed Target exit
def exit_condition_time_3(df, hold_bars=3):
    """Exit: Hold for N bars"""
    return hold_bars


def exit_condition_time_10(df, hold_bars=10):
    """Exit: Hold for N bars"""
    return hold_bars


def exit_condition_rsi_oversold(df, threshold=30):
    """Exit: RSI goes back below oversold level"""
    return df['RSI'] < threshold


def exit_condition_ma_cross(df, ma_col='MA_20'):
    """Exit: Price crosses below 20-day MA"""
    return df['Close'] < df[ma_col]


# Exit condition mapping
EXIT_CONDITIONS = {
    'Exit_RSI_55': {'func': exit_condition_rsi_55, 'params': {'threshold': 55}, 'type': 'rsi'},
    'Exit_RSI_60': {'func': exit_condition_rsi_60, 'params': {'threshold': 60}, 'type': 'rsi'},
    'Exit_RSI_65': {'func': exit_condition_rsi_65, 'params': {'threshold': 65}, 'type': 'rsi'},
    'Exit_RSI_70': {'func': exit_condition_rsi_70, 'params': {'threshold': 70}, 'type': 'rsi'},
    'Exit_RSI_75': {'func': exit_condition_rsi_75, 'params': {'threshold': 75}, 'type': 'rsi'},
    'Exit_Time_3': {'func': exit_condition_time_3, 'params': {'hold_bars': 3}, 'type': 'time'},
    'Exit_Time_10': {'func': exit_condition_time_10, 'params': {'hold_bars': 10}, 'type': 'time'},
    'Exit_RSI_Oversold': {'func': exit_condition_rsi_oversold, 'params': {'threshold': 30}, 'type': 'rsi_low'},
    'Exit_MA_Cross': {'func': exit_condition_ma_cross, 'params': {'ma_col': 'MA_20'}, 'type': 'ma'},
}


# Entry condition mapping
ENTRY_CONDITIONS = {
    'Entry_1': entry_condition_1,
    'Entry_2': entry_condition_2,
    'Entry_3': entry_condition_3,
    'Entry_4': entry_condition_4,
    'Entry_5': entry_condition_5,
    'Entry_6': entry_condition_6,
    'Entry_7': entry_condition_7,
    'Entry_8': entry_condition_8,
    'Entry_9': entry_condition_9,
    'Entry_10': entry_condition_10,
    'Entry_11': entry_condition_11,
    'Entry_12': entry_condition_12,
    'Entry_13': entry_condition_13,
    'Entry_14': entry_condition_14,
    'Entry_15': entry_condition_15,
}


def run_backtest(df, entry_name, exit_name):
    """
    Run backtest for a specific entry/exit combination.
    
    Returns:
        df_trades: DataFrame with trade signals
        equity_curve: Series with cumulative equity
    """
    entry_func = ENTRY_CONDITIONS[entry_name]
    exit_config = EXIT_CONDITIONS[exit_name]
    
    # Get entry signals
    entry_signal = entry_func(df).astype(int)
    
    # Initialize position (0 = flat, 1 = long)
    position = pd.Series(0, index=df.index)
    
    # Get exit parameters
    exit_type = exit_config['type']
    exit_params = exit_config['params']
    
    in_position = False
    entry_bar = 0
    bars_held = 0
    entry_price = 0
    
    for i in range(len(df)):
        if not in_position:
            # Try to enter on entry signal
            if entry_signal.iloc[i] == 1:
                position.iloc[i] = 1
                in_position = True
                entry_bar = i
                bars_held = 1
                entry_price = df['Close'].iloc[i]
        else:
            # Currently in position - check exit conditions
            should_exit = False
            
            if exit_type == 'time':
                # Time-based exit: hold for N bars
                hold_bars = exit_params.get('hold_bars', 5)
                if bars_held >= hold_bars:
                    should_exit = True
            elif exit_type == 'rsi':
                # RSI threshold exit
                threshold = exit_params.get('threshold', 65)
                if df['RSI'].iloc[i] > threshold:
                    should_exit = True
            elif exit_type == 'rsi_low':
                # RSI goes back to oversold (exit when momentum fades)
                threshold = exit_params.get('threshold', 30)
                if df['RSI'].iloc[i] < threshold:
                    should_exit = True
            elif exit_type == 'ma':
                # MA cross exit
                ma_col = exit_params.get('ma_col', 'MA_20')
                if df['Close'].iloc[i] < df[ma_col].iloc[i]:
                    should_exit = True
            
            if should_exit:
                position.iloc[i] = 0
                in_position = False
                bars_held = 0
                # Check if we can re-enter on same bar
                if entry_signal.iloc[i] == 1:
                    position.iloc[i] = 1
                    in_position = True
                    entry_bar = i
                    bars_held = 1
            else:
                position.iloc[i] = 1
                bars_held += 1
    
    # Calculate returns
    df['Position'] = position
    df['Strategy_Return'] = df['Position'].shift(1) * df['Daily_Return']
    df['Strategy_Return'] = df['Strategy_Return'].fillna(0)
    
    # Cumulative returns (equity curve)
    df['Cumulative_Strategy_Return'] = (1 + df['Strategy_Return']).cumprod()
    df['Cumulative_BuyAndHold'] = (1 + df['Daily_Return']).cumprod()
    
    # Extract trades
    trades = []
    in_trade = False
    for i in range(len(df)):
        if df['Position'].iloc[i] == 1 and not in_trade:
            # Entry
            trades.append({
                'Entry_Date': df['Date'].iloc[i],
                'Entry_Price': df['Close'].iloc[i],
                'Entry_RSI': df['RSI'].iloc[i],
                'Entry_IBS': df['IBS'].iloc[i],
            })
            in_trade = True
        elif df['Position'].iloc[i] == 0 and in_trade:
            # Exit
            trades[-1]['Exit_Date'] = df['Date'].iloc[i]
            trades[-1]['Exit_Price'] = df['Close'].iloc[i]
            trades[-1]['Exit_RSI'] = df['RSI'].iloc[i]
            trades[-1]['Return'] = (trades[-1]['Exit_Price'] / trades[-1]['Entry_Price']) - 1
            in_trade = False
    
    # Close any open position at the end
    if in_trade and len(trades) > 0:
        trades[-1]['Exit_Date'] = df['Date'].iloc[-1]
        trades[-1]['Exit_Price'] = df['Close'].iloc[-1]
        trades[-1]['Exit_RSI'] = df['RSI'].iloc[-1]
        trades[-1]['Return'] = (trades[-1]['Exit_Price'] / trades[-1]['Entry_Price']) - 1
    
    df_trades = pd.DataFrame(trades)
    
    return df, df_trades


def calculate_metrics(df_trades, df_full):
    """Calculate performance metrics"""
    if len(df_trades) == 0:
        return {
            'Num_Trades': 0,
            'Total_Return': 0,
            'Profit_Factor': 0,
            'Win_Rate': 0,
        }
    
    # Calculate returns
    trades_returns = df_trades['Return'].dropna()
    
    if len(trades_returns) == 0:
        return {
            'Num_Trades': 0,
            'Total_Return': 0,
            'Profit_Factor': 0,
            'Win_Rate': 0,
        }
    
    wins = trades_returns[trades_returns > 0]
    losses = trades_returns[trades_returns < 0]
    
    total_wins = wins.sum() if len(wins) > 0 else 0
    total_losses = abs(losses.sum()) if len(losses) > 0 else 0
    
    profit_factor = total_wins / total_losses if total_losses > 0 else float('inf') if total_wins > 0 else 0
    
    win_rate = len(wins) / len(trades_returns) * 100 if len(trades_returns) > 0 else 0
    
    # Total return from equity curve
    final_return = df_full['Cumulative_Strategy_Return'].iloc[-1] - 1
    
    return {
        'Num_Trades': len(trades_returns),
        'Total_Return': final_return * 100,  # As percentage
        'Profit_Factor': profit_factor,
        'Win_Rate': win_rate,
    }


def run_all_strategies(df):
    """Run all 45 strategy combinations (5 top entries x 9 exits)"""
    results = []
    equity_curves = {}
    dates = df['Date']  # Store dates for plotting
    
    # Top 5 entries from prior analysis (all with FTD criteria)
    top_entries = ['Entry_10', 'Entry_15', 'Entry_7', 'Entry_8', 'Entry_5']
    
    # All 9 exit variations (removed Target)
    exit_names = list(EXIT_CONDITIONS.keys())
    
    for entry_name in top_entries:
        for exit_name in exit_names:
            strategy_name = f"{entry_name}_{exit_name}"
            print(f"Running {strategy_name}...")
            
            df_result, df_trades = run_backtest(df, entry_name, exit_name)
            metrics = calculate_metrics(df_trades, df_result)
            metrics['Strategy'] = strategy_name
            
            results.append(metrics)
            # Store equity curve with dates
            equity_curves[strategy_name] = pd.Series(
                df_result['Cumulative_Strategy_Return'].values, 
                index=dates
            )
            
            print(f"  Trades: {metrics['Num_Trades']}, Return: {metrics['Total_Return']:.2f}%, PF: {metrics['Profit_Factor']:.2f}")
    
    return results, equity_curves


def create_performance_table(results):
    """Create a formatted performance table"""
    df_results = pd.DataFrame(results)
    df_results = df_results[['Strategy', 'Num_Trades', 'Total_Return', 'Profit_Factor', 'Win_Rate']]
    df_results = df_results.sort_values('Total_Return', ascending=False)
    
    # Format columns
    df_results['Total_Return'] = df_results['Total_Return'].apply(lambda x: f"{x:.2f}%")
    df_results['Profit_Factor'] = df_results['Profit_Factor'].apply(lambda x: f"{x:.2f}")
    df_results['Win_Rate'] = df_results['Win_Rate'].apply(lambda x: f"{x:.1f}%")
    
    return df_results


def plot_performance(equity_curves, results, save_path='ftd_performance.png'):
    """Create performance plot with log scale and date x-axis - top 5 by profit factor"""
    # Get results and filter for meaningful trades (>= 10 trades)
    df_results = pd.DataFrame(results)
    df_filtered = df_results[df_results['Num_Trades'] >= 10].copy()
    
    # Replace inf with a large number for sorting
    df_filtered['Profit_Factor'] = df_filtered['Profit_Factor'].replace(float('inf'), 999)
    
    # Get top 5 strategies by profit factor
    top_5 = df_filtered.nlargest(5, 'Profit_Factor')['Strategy'].tolist()
    
    plt.figure(figsize=(14, 8))
    
    # Use log scale for y-axis
    plt.yscale('log')
    
    # Plot only top 5 strategies
    for name, curve in equity_curves.items():
        if name in top_5:
            plt.plot(curve.index, curve.values, label=name, linewidth=1.5)
    
    plt.xlabel('Date', fontsize=12)
    plt.ylabel('Cumulative Return (Log Scale)', fontsize=12)
    plt.title('FTD Strategy Performance - Top 5 by Profit Factor', fontsize=14)
    plt.legend(loc='upper left', fontsize=10)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"Performance plot saved to {save_path}")
    print(f"Top 5 strategies by Profit Factor (min 10 trades): {top_5}")


def create_rules_table():
    """Create a table of all strategy rules for the 50 strategies"""
    # Top 5 entries (all with FTD criteria)
    entry_rules = [
        {'Strategy': 'Entry_10', 'Description': 'RSI < 30 + FTD + Price above 200-day MA (Trend Confirmation)', 'Category': 'FTD Entry'},
        {'Strategy': 'Entry_15', 'Description': 'RSI < 30 + FTD + Large green bar (>1.5%) (Wide Bar)', 'Category': 'FTD Entry'},
        {'Strategy': 'Entry_7', 'Description': 'RSI < 30 + FTD + Volume > 1.5x average (Volume Spike)', 'Category': 'FTD Entry'},
        {'Strategy': 'Entry_8', 'Description': 'RSI < 30 + FTD + Gap up > 0.5% (Gap Up)', 'Category': 'FTD Entry'},
        {'Strategy': 'Entry_5', 'Description': 'RSI < 40 + FTD (IBS > 0.5, any positive return)', 'Category': 'FTD Entry'},
    ]
    
    # Exit rules - RSI threshold variations (5)
    rsi_exit_rules = [
        {'Strategy': 'Exit_RSI_55', 'Description': 'Exit when RSI > 55', 'Category': 'RSI Threshold'},
        {'Strategy': 'Exit_RSI_60', 'Description': 'Exit when RSI > 60', 'Category': 'RSI Threshold'},
        {'Strategy': 'Exit_RSI_65', 'Description': 'Exit when RSI > 65', 'Category': 'RSI Threshold'},
        {'Strategy': 'Exit_RSI_70', 'Description': 'Exit when RSI > 70', 'Category': 'RSI Threshold'},
        {'Strategy': 'Exit_RSI_75', 'Description': 'Exit when RSI > 75', 'Category': 'RSI Threshold'},
    ]
    
    # Exit rules - new methodologies (4 - removed Target)
    new_exit_rules = [
        {'Strategy': 'Exit_Time_3', 'Description': 'Hold for 3 bars then exit', 'Category': 'Time-based'},
        {'Strategy': 'Exit_Time_10', 'Description': 'Hold for 10 bars then exit', 'Category': 'Time-based'},
        {'Strategy': 'Exit_RSI_Oversold', 'Description': 'Exit when RSI goes back below 30 (momentum fades)', 'Category': 'RSI Reversal'},
        {'Strategy': 'Exit_MA_Cross', 'Description': 'Exit when price crosses below 20-day MA', 'Category': 'MA Cross'},
    ]
    
    return pd.DataFrame(entry_rules), pd.DataFrame(rsi_exit_rules + new_exit_rules)


def save_to_excel(df_results, entry_rules, exit_rules, filename='ftd_table2.xlsx'):
    """Save performance and rules to Excel with formatting"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write performance sheet (sorted by profit factor descending)
        df_results_sorted = df_results.sort_values('Profit_Factor', ascending=False)
        df_results_sorted.to_excel(writer, sheet_name='Perf', index=False)
        
        # Write entry rules sheet
        entry_rules.to_excel(writer, sheet_name='Entries', index=False)
        
        # Write exit rules sheet
        exit_rules.to_excel(writer, sheet_name='Exits', index=False)
        
        # Get workbook and worksheets
        workbook = writer.book
        perf_sheet = writer.sheets['Perf']
        entry_sheet = writer.sheets['Entries']
        exit_sheet = writer.sheets['Exits']
        
        # Format Performance sheet
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in perf_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adjust column widths
        perf_sheet.column_dimensions['A'].width = 30  # Strategy
        perf_sheet.column_dimensions['B'].width = 12  # Num_Trades
        perf_sheet.column_dimensions['C'].width = 14  # Total_Return
        perf_sheet.column_dimensions['D'].width = 14  # Profit_Factor
        perf_sheet.column_dimensions['E'].width = 12  # Win_Rate
        
        # Format Entry Rules sheet
        for cell in entry_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        entry_sheet.column_dimensions['A'].width = 15
        entry_sheet.column_dimensions['B'].width = 75
        entry_sheet.column_dimensions['C'].width = 15
        
        # Format Exit Rules sheet
        for cell in exit_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        exit_sheet.column_dimensions['A'].width = 20
        exit_sheet.column_dimensions['B'].width = 60
        exit_sheet.column_dimensions['C'].width = 18
        
        # Add borders to cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in perf_sheet.iter_rows(min_row=2, max_row=perf_sheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
        
        for row in entry_sheet.iter_rows(min_row=2, max_row=entry_sheet.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
        
        for row in exit_sheet.iter_rows(min_row=2, max_row=exit_sheet.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
    
    print(f"Excel file saved to {filename}")


def main():
    print("=" * 60)
    print("FTD Strategy - Enhanced Exit Analysis (45 strategies)")
    print("=" * 60)
    
    # Download data
    df = download_spy_data()
    
    # Run all strategies
    results, equity_curves = run_all_strategies(df)
    
    # Create and display performance table (sorted by profit factor)
    df_results = create_performance_table(results)
    df_results_by_pf = df_results.sort_values('Profit_Factor', ascending=False)
    print("\n" + "=" * 60)
    print("Performance Table (sorted by Profit Factor)")
    print("=" * 60)
    print(df_results_by_pf.to_string(index=False))
    
    # Create rules tables
    entry_rules, exit_rules = create_rules_table()
    
    # Save to Excel (ftd_table2.xlsx)
    save_to_excel(df_results, entry_rules, exit_rules, 'ftd_table2.xlsx')
    
    # Create performance plot (top 5 by profit factor) - ftd_performance2.png
    plot_performance(equity_curves, results, 'ftd_performance2.png')
    
    print("\n" + "=" * 60)
    print("Analysis Complete!")
    print("=" * 60)
    
    return df_results, equity_curves


if __name__ == "__main__":
    df_results, equity_curves = main()